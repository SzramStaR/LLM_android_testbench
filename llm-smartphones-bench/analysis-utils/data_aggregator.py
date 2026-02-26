import copy
import os
import json
import statistics
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def transform_mlc_model_name(model_name):
    if model_name.endswith("-MLC"):
        model_name = model_name[:-4]

    if "-Instruct-" in model_name:
        model_name = model_name.replace("-Instruct-", "-")

    parts = model_name.split("-")

    if parts and "_" in parts[-1]:
        quantization = parts[-1]
        model_parts = parts[:-1]
        transformed_model = " ".join(model_parts)
        result = f"{transformed_model} {quantization}"
    else:
        result = model_name.replace("-", " ")

    return result


def read_perf_data(directory_path):
    parsed_json_data = []

    for filename in os.listdir(directory_path):
        if filename.endswith(".json"):
            file_path = os.path.join(directory_path, filename)
            with open(file_path, "r", encoding="utf-8") as f:
                try:
                    data = json.load(f)

                    data["runId"] = filename.replace(".json", "").replace(
                        "benchmark-", ""
                    )

                    for key in ["data", "phoneData"]:
                        if key in data and isinstance(data[key], str):
                            try:
                                data[key] = json.loads(data[key])
                            except json.JSONDecodeError as e:
                                print(f"Could not parse {key} in {filename}: {e}")

                    parsed_json_data.append(data)
                except json.JSONDecodeError as e:
                    print(f"Error decoding JSON in file {filename}: {e}")

    return parsed_json_data


def read_and_postprocess_perf_data(rootFolder: str) -> list[dict]:
    postprocessed_run_data = []
    all_data = read_perf_data(rootFolder)
    for item in all_data:
        data = item["data"]

        version = item.get("version", "")
        if version.startswith("ExecutorTorch"):
            application = "ExecuTorch"
        elif version.startswith("MLC"):
            application = "MLC"
        else:
            application = "llama.cpp"

        if application in ["ExecuTorch", "MLC"]:
            family = "Llama 3.1 8B" if "3.1" in item["model"] else "Llama 3.2 3B"
        else:
            family = item["family"]

        model = item["model"]
        model_replacements = {
            "llama3_2_3B_spinquant.pte": "Llama 3.2 3B SpinQuant",
            "llama3_2_3B_bf16.pte": "Llama 3.2 3B BF16",
            "llama3_1_8b_spinquant.pte": "Llama 3.1 8B SpinQuant",
        }

        model = model_replacements.get(model, model)

        if application == "MLC":
            model = transform_mlc_model_name(model)

        flatten_run_base = {
            "application": application,
            "runId": item["runId"],
            "userId": item["userId"],
            "model": model,
            "family": family,
            "deviceModel": item["phoneData"].get(
                "deviceModel", item["phoneData"].get("model", "UKNW")
            ),
            "deviceBrand": item["phoneData"].get(
                "deviceBrand", item["phoneData"].get("brand", "UKNW")
            ),
            "systemName": item["phoneData"]["systemName"],
            "systemVersion": item["phoneData"]["systemVersion"],
            "totalMemory": item["phoneData"]["totalMemory"],
        }
        for runNum, run in enumerate(data["runs"]):
            flatten_run = copy.deepcopy(flatten_run_base)

            new_sensorTempretures = []
            for sensorTempreture in run["sensorTempreratures"]:
                new_sensorTempreture = [
                    x
                    for x in sensorTempreture.split("\n")
                    if x != "" and "273000" not in x
                ]
                new_sensorTempreture = {
                    item.split(": ")[0]: float(item.split(": ")[1])
                    for item in new_sensorTempreture
                }
                new_sensorTempretures.append(new_sensorTempreture)
            run["sensorTempreratures"] = new_sensorTempretures

            ram_values = run["ram"]
            ram_max = max(ram_values)
            threshold = 0.7 * ram_max
            ram_filtered = [value for value in ram_values if value >= threshold]
            if not ram_filtered:
                ram_filtered = ram_values

            flatten_run["runNum"] = runNum + 1
            flatten_run["ramMax"] = ram_max
            flatten_run["ramMin"] = min(ram_values)
            flatten_run["ramMedian"] = statistics.median(ram_filtered)
            flatten_run["ramAvg"] = statistics.mean(ram_filtered)
            flatten_run["ramStd"] = statistics.stdev(ram_values)
            flatten_run["inputTokens"] = run["inputTokens"]
            flatten_run["outputTokens"] = run["outputTokens"]
            flatten_run["tps"] = run["tps"]
            flatten_run["ttft"] = run["ttft"]
            flatten_run["prefillSpeed"] = run["inputTokens"] / (run["ttft"] / 1000.0)
            flatten_run["inferenceTime"] = run["inferenceTime"]
            flatten_run["battery_pct_bef"] = run["battery"][0]
            flatten_run["battery_pct_aft"] = run["battery"][1]
            flatten_run["battery_temp_bef"] = run["batteryTempreture"][0]
            flatten_run["battery_temp_aft"] = run["batteryTempreture"][1]

            skip_keys = ["isCharging", "voltageV", "currentDrawMa"]

            if "batteryInfos" in run:
                for x1, x2 in zip(
                    run["batteryInfos"][0].items(), run["batteryInfos"][1].items()
                ):
                    if x1[0] not in skip_keys:
                        flatten_run[f"batInfo_{x1[0]}_bef"] = x1[1]
                    if x2[0] not in skip_keys:
                        flatten_run[f"batInfo_{x2[0]}_aft"] = x2[1]

            if "batteryInfos" in run:
                for x1, x2 in zip(
                    run["sensorTempreratures"][0].items(),
                    run["sensorTempreratures"][1].items(),
                ):
                    flatten_run[f"sensTemp_{x1[0]}_bef"] = x1[1]
                    flatten_run[f"sensTemp_{x2[0]}_aft"] = x2[1]

            postprocessed_run_data.append(flatten_run)

    return postprocessed_run_data


def sort_by_quant(x):
    quant = x["model"].split(" ")[2]
    if "I" in quant:
        quant = quant.replace("I", "") + "I"
    return quant


def read_eval_data(directory_path):
    parsed_json_data = []

    for filename in os.listdir(directory_path):
        if filename.endswith(".json"):
            file_path = os.path.join(directory_path, filename)
            with open(file_path, "r", encoding="utf-8") as f:
                try:
                    data = json.load(f)
                    parsed_data = {}

                    if "MMLU" in filename or "IFEval" in filename:
                        for key, value in data.items():
                            if value == "N/A":
                                continue
                            if key.endswith("acc") or key.endswith("acc_stderr"):
                                parsed_data[key] = value
                    else:
                        continue

                    parsed_data = dict(sorted(parsed_data.items()))

                    model, _, quant = filename.replace(".json", "").split("-")

                    parsed_data = {
                        "model": model.replace("_", " ") + " " + quant,
                        **parsed_data,
                    }

                    parsed_json_data.append(parsed_data)
                except json.JSONDecodeError as e:
                    print(f"Error decoding JSON in file {filename}: {e}")

    parsed_json_data.sort(key=sort_by_quant)

    return parsed_json_data


def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_value = str(cell.value)
                max_length = max(max_length, len(cell_value))
            except (AttributeError, TypeError):
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


llama_offline_perf_data = read_and_postprocess_perf_data("./data/llama31_8b_perf")
llama_offline_perf_data.sort(
    key=lambda x: (
        x["application"],
        x["deviceModel"],
        x["family"],
        x["model"],
        x["inputTokens"],
    )
)

llama31_mmlu_data = read_eval_data("./data/llama31_8b_mmlu")
llama31_ifeval_data = read_eval_data("./data/llama31_8b_ifeval")
llama32_mmlu_data = read_eval_data("./data/llama32_3b_mmlu")
llama32_ifeval_data = read_eval_data("./data/llama32_3b_ifeval")

file_name = "benchmark_results.xlsx"
wb = Workbook()

default_sheet = wb.active
wb.remove(default_sheet)

data_batches = {
    "llama_perf": llama_offline_perf_data,
    "llama31_mmlu": llama31_mmlu_data,
    "llama31_ifeval": llama31_ifeval_data,
    "llama32_mmlu": llama32_mmlu_data,
    "llama32_ifeval": llama32_ifeval_data,
}

for sheet_name, data in data_batches.items():
    df = pd.DataFrame(data)
    ws = wb.create_sheet(title=sheet_name)

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(df.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    adjust_column_width(ws)

wb.save(file_name)
print(f"Excel file '{file_name}' created successfully with multiple worksheets.")
